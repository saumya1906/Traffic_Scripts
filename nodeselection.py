#!/usr/bin/env python
import json
from urllib.request import urlopen
import requests
import xlrd
import math
from apscheduler.schedulers.blocking import BlockingScheduler
from xlutils.copy import copy    
from xlrd import open_workbook
import pandas as pd
import openpyxl
import datetime
import os
import pathlib

def prog():
	today = datetime.date.today()
	workbook1 = xlrd.open_workbook('nodes.xlsx', on_demand=True)
	sheet1 = workbook1.sheet_by_index(0)
	arrayofnodelat = sheet1.col_values(0)
	arrayofnodelong = sheet1.col_values(1)
	nearlat = sheet1.col_values(2)
	nearlong = sheet1.col_values(3)
	speedval1=[]
	# speedval2 = []
	# speedval3 = []
	speedval1_up = []
	# speedval2_up = []
	# speedval3_up = []
	Node = []
	Time =[]
	for i in range(1,14):
		Node.append(i)
	url = 'https://maps.googleapis.com/maps/api/distancematrix/json?origins='+str(arrayofnodelat[1])+','+str(arrayofnodelong[1])+'&destinations='+str(nearlat[1])+','+str(nearlong[1])+'&departure_time=now&traffic_model=best_guess&mode=driving&key=AIzaSyDI-v-cqXgQepTqi2NrejoyWNsw2cdFFA0'
	res=requests.get(url)
	result1=res.json()
	with open('419down.json','a') as fp:
		json.dump(result1,fp)
		fp.close()
	# print(result1)
	a = result1['rows'][0]['elements'][0]['duration_in_traffic']['value']
	time=float(a)
	a = result1['rows'][0]['elements'][0]['distance']['value']
	dist = float(a)
	speed = dist/(time + 1)
	speedval1.append(speed)
	
	time2 = str(datetime.datetime.now().time())
	# speedval3.append(speed)
	Time.append(time2)
	for i in range(2,len(arrayofnodelong)):
		url = 'https://maps.googleapis.com/maps/api/distancematrix/json?origins='+str(arrayofnodelat[i])+','+str(arrayofnodelong[i])+'&destinations='+str(arrayofnodelat[i-1])+','+str(arrayofnodelong[i-1])+'&departure_time=now&traffic_model=best_guess&mode=driving&key=AIzaSyDI-v-cqXgQepTqi2NrejoyWNsw2cdFFA0'
		res=requests.get(url)
		result1=res.json()
		with open('419down.json','a') as fp:
			json.dump(result1,fp)
			fp.close()
		#print(result1)
		a = result1['rows'][0]['elements'][0]['duration_in_traffic']['value']
		time=float(a)
		a = result1['rows'][0]['elements'][0]['distance']['value']
		dist = float(a)
		speed = dist/(time + 1)
		#time2 = str(datetime.datetime.now().time())
		speedval1.append(speed)
		
		time2 = str(datetime.datetime.now().time())
		# speedval3.append(speed)
		Time.append(time2)
		
	for i in range(1,len(arrayofnodelong)-1):
		url = 'https://maps.googleapis.com/maps/api/distancematrix/json?origins='+str(arrayofnodelat[i])+','+str(arrayofnodelong[i])+'&destinations='+str(arrayofnodelat[i+1])+','+str(arrayofnodelong[i+1])+'&departure_time=now&traffic_model=best_guess&mode=driving&key=AIzaSyDI-v-cqXgQepTqi2NrejoyWNsw2cdFFA0'
		res=requests.get(url)
		result1=res.json()
		with open('419up.json','a') as fp:
			json.dump(result1,fp)
			fp.close()
		#print(result1)
		a = result1['rows'][0]['elements'][0]['duration_in_traffic']['value']
		time=float(a)
		a = result1['rows'][0]['elements'][0]['distance']['value']
		dist = float(a)
		speed = dist/(time + 1)
		#time2 = str(datetime.datetime.now().time())
		speedval1_up.append(speed)
		# time2 = str(datetime.datetime.now().time())
		# # speedval3.append(speed)
		# Time.append(time2)
	
	url = 'https://maps.googleapis.com/maps/api/distancematrix/json?origins='+str(arrayofnodelat[13])+','+str(arrayofnodelong[13])+'&destinations='+str(nearlat[13])+','+str(nearlong[13])+'&departure_time=now&traffic_model=best_guess&mode=driving&key=AIzaSyDI-v-cqXgQepTqi2NrejoyWNsw2cdFFA0'
	res=requests.get(url)
	result1=res.json()
	with open('419up.json','a') as fp:
		json.dump(result1,fp)
		fp.close()
	#print(result1)
	a = result1['rows'][0]['elements'][0]['duration_in_traffic']['value']
	time=float(a)
	a = result1['rows'][0]['elements'][0]['distance']['value']
	dist = float(a)
	speed = dist/(time + 1)
	speedval1_up.append(speed)
	# time2 = str(datetime.datetime.now().time())
	# 	# speedval3.append(speed)
	# Time.append(time2)
	
	path = pathlib.Path('419_googleapi_down'+str(today)+'.xlsx')
	print(len(Time))
	print(len(speedval1))
	print(len(speedval1_up))
	if(path.is_file()==False):
		#fo.close()
		#print("hello2")
		df = pd.DataFrame({'Node':Node,'Best Guess Speed':speedval1, 'Time':Time})
		book_ro = pd.ExcelWriter('419_googleapi_down'+str(today)+'.xlsx', engine='xlsxwriter')
		df.to_excel(book_ro, sheet_name='Sheet1')
		book_ro.save()
		df = pd.DataFrame({'Node':Node,'Best Guess Speed':speedval1_up, 'Time':Time})
		book_ro = pd.ExcelWriter('419_googleapi_up'+str(today)+'.xlsx', engine='xlsxwriter')
		df.to_excel(book_ro, sheet_name='Sheet1')
		book_ro.save()

	else:
		#fo.close()
		#print("hello")
		book_ro = openpyxl.load_workbook('419_googleapi_down'+str(today)+'.xlsx')
		#book = copy(book_ro)  # creates a writeable copy
		sheet1 = book_ro.active # get a first sheet
		#max1 = sheet1.max_row
		for i in range(0,13):
			sheet1.append({3:Node[i],2:speedval1[i], 4:Time[i]})

		book_ro.save('419_googleapi_down'+str(today)+'.xlsx')
		book_ro = openpyxl.load_workbook('419_googleapi_up'+str(today)+'.xlsx')
		#book = copy(book_ro)  # creates a writeable copy
		sheet1 = book_ro.active # get a first sheet
		#max1 = sheet1.max_row
		for i in range(0,13):
			sheet1.append({3:Node[i],2:speedval1_up[i],4:Time[i]})

		book_ro.save('419_googleapi_up'+str(today)+'.xlsx')

scheduler = BlockingScheduler()
scheduler.add_job(prog, 'cron', hour='*', minute='0')
scheduler.start()
